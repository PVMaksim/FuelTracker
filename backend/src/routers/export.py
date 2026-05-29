"""Export endpoints: CSV and Excel."""
import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, asc
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.connection import get_db
from src.database.models import Refuel
from src.config import settings

router = APIRouter(tags=["export"])

COLUMNS = ["Дата", "Пробег (км)", "Пробег за интервал (км)", "Цена за литр (₽)",
           "Сумма (₽)", "Литры", "Расход (л/100км)", "₽/км", "Тип топлива", "Заметки"]


async def verify_key(api_key: str = Query(alias="api_key")):
    """Проверка ключа через query-параметр (для скачивания файлов)."""
    from fastapi import HTTPException
    if api_key != settings.api_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    return api_key


@router.get("/refuels/export/csv")
async def export_csv(
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_key),
    car_id: str | None = None,
):
    query = select(Refuel).order_by(asc(Refuel.created_at))
    if car_id:
        query = query.where(Refuel.car_id == car_id)
    rows = await db.execute(query)
    refuels = rows.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(COLUMNS)
    for r in refuels:
        writer.writerow([
            r.created_at.strftime("%d.%m.%Y %H:%M"),
            r.odometer, r.distance or "",
            f"{float(r.fuel_price):.2f}", f"{float(r.total_cost):.2f}",
            f"{float(r.liters):.2f}",
            f"{float(r.consumption):.2f}" if r.consumption else "",
            f"{float(r.cost_per_km):.2f}" if r.cost_per_km else "",
            r.fuel_type or "", r.notes or "",
        ])
    output.seek(0)
    filename = f"fueltracker_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/refuels/export/xlsx")
async def export_xlsx(
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_key),
    car_id: str | None = None,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(Refuel).order_by(asc(Refuel.created_at))
    if car_id:
        query = query.where(Refuel.car_id == car_id)
    rows = await db.execute(query)
    refuels = rows.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заправки"
    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in refuels:
        ws.append([
            r.created_at.strftime("%d.%m.%Y %H:%M"),
            r.odometer, r.distance or "",
            float(r.fuel_price), float(r.total_cost),
            float(r.liters),
            float(r.consumption) if r.consumption else "",
            float(r.cost_per_km) if r.cost_per_km else "",
            r.fuel_type or "", r.notes or "",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"fueltracker_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
